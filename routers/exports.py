import io
from datetime import datetime, timedelta, timezone
from typing import Optional
from urllib.parse import quote

KST = timezone(timedelta(hours=9))

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy import or_
from sqlalchemy.orm import Session

import models
from database import get_db
from routers.auth import get_current_user
from routers.deliveries import _apply_visibility_filter

router = APIRouter()

STATUS_LABELS = {
    "wait":     "대기중",
    "start":    "업무시작",
    "loaded":   "상차",
    "unloaded": "하차",
    "weighed":  "계근표 등록",
    "done":     "완료",
    "cancel":   "취소",
}


# ── 엑셀 내보내기 ──────────────────────────────────────────────────────────────
@router.get("/excel")
def export_excel(
    status: Optional[str] = None,
    driver_id: Optional[int] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    delivery_type: Optional[str] = None,
    company: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    query = _apply_visibility_filter(db.query(models.Delivery), db, current_user)
    if status:
        query = query.filter(models.Delivery.status == status)
    if company:
        query = query.filter(models.Delivery.company == company)
    if delivery_type:
        # 화면 목록과 같은 기준으로 거른다 (빈 값은 출하로 취급)
        if delivery_type == "출하":
            query = query.filter(or_(models.Delivery.delivery_type == "출하",
                                     models.Delivery.delivery_type.is_(None),
                                     models.Delivery.delivery_type == ""))
        else:
            query = query.filter(models.Delivery.delivery_type == delivery_type)
    if driver_id:
        query = query.filter(models.Delivery.driver_id == driver_id)
    if date_from:
        query = query.filter(models.Delivery.scheduled_date >= date_from)
    if date_to:
        query = query.filter(models.Delivery.scheduled_date <= date_to)
    deliveries = query.order_by(
        models.Delivery.scheduled_date, models.Delivery.delivery_time
    ).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "배송 목록"

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # 제목 (17열 전체 병합)
    ws.merge_cells("A1:Q1")
    ws["A1"] = "탱크로리 배송 내역"
    ws["A1"].font = Font(bold=True, size=15, color="1E3A5F")
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 34

    ws.merge_cells("A2:Q2")
    ws["A2"] = f"출력일: {datetime.now(KST).strftime('%Y년 %m월 %d일 %H:%M')}"
    ws["A2"].font = Font(size=10, color="6B7280")
    ws["A2"].alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[2].height = 18

    # 헤더 (17열)
    headers = [
        "번호", "유형", "업체명", "목적지", "품목", "수량(Kg)",
        "기사명", "차량번호", "배송날짜", "배송시간",
        "업무시작", "상차", "하차", "계근표등록", "완료시간", "상태", "특이사항",
    ]
    col_widths = [10, 8, 16, 22, 20, 10, 10, 13, 13, 10, 10, 10, 10, 11, 10, 10, 26]

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = w
    ws.row_dimensions[3].height = 22

    status_fills = {
        "wait":     PatternFill("solid", fgColor="FEF3C7"),
        "start":    PatternFill("solid", fgColor="DBEAFE"),
        "loaded":   PatternFill("solid", fgColor="EDE9FE"),
        "unloaded": PatternFill("solid", fgColor="CCFBF1"),
        "done":     PatternFill("solid", fgColor="DCFCE7"),
        "cancel":   PatternFill("solid", fgColor="FEE2E2"),
    }

    for row_idx, d in enumerate(deliveries, 4):
        row_values = [
            # 화면·푸시 알림과 같은 배송번호. 예전에는 줄 번호(1,2,3…)라
            # 거르는 조건에 따라 매번 달라져 어느 배송건인지 알 수 없었다.
            f"D{d.id:03d}",
            d.delivery_type or "출하",
            d.company,
            d.destination,
            d.item_name,
            d.quantity,
            d.driver_user.name if d.driver_user else "",
            d.vehicle_number or "",
            d.scheduled_date,
            d.delivery_time,
            d.work_start_time or "-",
            d.loading_complete_time or "-",
            d.unloaded_time or "-",
            d.weighed_time or "-",
            d.complete_time or "-",
            STATUS_LABELS.get(d.status, d.status),
            d.notes or "",
        ]
        status_fill = status_fills.get(d.status)
        for col_idx, val in enumerate(row_values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            # 17열(특이사항)만 왼쪽 정렬, 16열(상태)에 상태 색상
            cell.alignment = left_wrap if col_idx == 17 else center
            cell.border = thin_border
            if col_idx == 16 and status_fill:
                cell.fill = status_fill
        ws.row_dimensions[row_idx].height = 18

    ws.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"배송목록_{datetime.now(KST).strftime('%Y%m%d')}.xlsx"
    encoded = quote(filename)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
    )


# ── 계근표 사진의 '원본 파일명 → 고객사' 대조표 ─────────────────────────────────
# 구글 시트에 이미 쌓인 행의 '거래처'를 배송카드의 고객사명으로 채우기 위한 것.
# 시트의 '파일명' 열에는 업로드 당시의 원본 이름(IMG_2847.jpg 등)이 들어 있는데,
# 그 이름만으로는 어느 배송건인지 알 수 없어 앱 DB에서 뽑아 준다.
# 같은 원본 이름이 서로 다른 고객사의 배송건에 여러 번 쓰였을 수 있으므로,
# 그런 이름은 '중복' 으로 표시해 잘못 채우지 않도록 한다.
@router.get("/photo-company-map")
def photo_company_map(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    if current_user.role != "superadmin":
        from fastapi import HTTPException
        raise HTTPException(status_code=403, detail="슈퍼관리자만 내려받을 수 있습니다.")

    rows = (
        db.query(models.DeliveryPhoto.filename, models.Delivery.company, models.Delivery.id)
        .join(models.Delivery, models.DeliveryPhoto.delivery_id == models.Delivery.id)
        .filter(models.DeliveryPhoto.filename.isnot(None))
        .all()
    )

    # 원본 이름 하나가 어느 고객사들에 걸쳐 쓰였는지 모은다
    by_name = {}
    for fname, company, did in rows:
        if not fname:
            continue
        by_name.setdefault(fname, []).append((company or "", did))

    buf = io.StringIO()
    buf.write("\ufeff")                      # 엑셀·시트에서 한글이 깨지지 않도록
    buf.write("원본파일명,고객사,배송번호,중복\n")

    def esc(v):
        v = str(v or "")
        return '"' + v.replace('"', '""') + '"' if any(c in v for c in ',"\n') else v

    for fname, entries in sorted(by_name.items()):
        companies = {c for c, _ in entries}
        dup = "중복" if len(companies) > 1 else ""
        company, did = entries[0]
        buf.write(f"{esc(fname)},{esc(company)},D{did:03d},{dup}\n")

    data = buf.getvalue().encode("utf-8")
    filename = f"계근표_파일명_고객사_대조표_{datetime.now(KST).strftime('%Y%m%d')}.csv"
    return StreamingResponse(
        io.BytesIO(data),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )
